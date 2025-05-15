import asyncio  # pour gérer les coroutines asynchrones
import csv  # pour écrire dans un fichier CSV
import os  # pour gérer les dossiers
import random  # pour générer des chaînes aléatoires
import re  # pour nettoyer les noms de fichiers
import string  # pour les caractères des chaînes aléatoires
import time  # pour mesurer le temps d'exécution
from typing import Dict, List  # pour les annotations de type
from urllib.parse import urljoin  # pour résoudre les URLs relatives

import aiohttp  # pour les requêtes HTTP asynchrones
import matplotlib.pyplot as plt  # pour les visualisations
import pandas as pd  # pour le traitement des données
import requests  # pour les requêtes HTTP synchrones (étape 1)
from bs4 import BeautifulSoup  # pour parser le HTML
from openpyxl import Workbook  # pour créer un fichier Excel
from openpyxl.styles import PatternFill  # pour colorer les cellules
from openpyxl.utils import get_column_letter  # pour obtenir la lettre de la colonne


def fetch_page(url: str) -> BeautifulSoup:
    """Récupère le contenu HTML de la page à l'URL donnée (synchrone, pour l'étape 1).
    Args:
        url: L'URL de la page à scraper.
    Returns:
        Un objet BeautifulSoup contenant le contenu HTML parsé.
    """
    response = requests.get(url)
    response.raise_for_status()  # exception si la requête échoue
    return BeautifulSoup(response.text, 'html.parser')


async def async_fetch_page(session: aiohttp.ClientSession, url: str) -> BeautifulSoup:
    """Récupère le contenu HTML de la page à l'URL donnée (asynchrone).
    Args:
        session: Session aiohttp pour les requêtes.
        url: L'URL de la page à scraper.
    Returns:
        Un objet BeautifulSoup contenant le contenu HTML parsé.
    """
    async with session.get(url) as response:
        response.raise_for_status()
        text = await response.text()
        return BeautifulSoup(text, 'html.parser')


def extract_book_info(soup: BeautifulSoup, url: str) -> Dict[str, str]:
    """Extrait les informations du livre depuis la soupe BeautifulSoup.
    Args:
        soup: Objet BeautifulSoup contenant le HTML parsé.
        url: L'URL de la page du livre.
    Returns:
        Un dictionnaire contenant les informations du livre.
    """
    # UPC
    upc = soup.find('th', string='UPC').find_next('td').text

    # Titre
    title = soup.find('h1').text

    # Prix avec taxes
    price_including_tax = soup.find('th', string='Price (incl. tax)').find_next('td').text

    # Prix chacun des livres
    price_excluding_tax = soup.find('th', string='Price (excl. tax)').find_next('td').text

    # Nombre disponible
    number_available_text = soup.find('th', string='Availability').find_next('td').text
    number_available = ''.join(filter(str.isdigit, number_available_text))  # Extrait le nombre

    # Description du produit
    product_description = soup.find('div', id='product_description').find_next('p').text

    # Catégorie
    category = soup.find('ul', class_='breadcrumb').find_all('li')[2].text.strip()

    # Note (review rating)
    rating_class = soup.find('p', class_='star-rating')['class'][1]
    rating_map = {'One': 1, 'Two': 2, 'Three': 3, 'Four': 4, 'Five': 5}
    review_rating = str(rating_map[rating_class])  # Converti en string

    # URL de l'image
    image_url = soup.find('div', class_='item active').find('img')['src']
    image_url = 'http://books.toscrape.com' + image_url.replace('../..', '')

    return {
        'product_page_url': url,
        'universal_product_code': upc,
        'title': title,
        'price_including_tax': price_including_tax,
        'price_excluding_tax': price_excluding_tax,
        'number_available': number_available,
        'product_description': product_description,
        'category': category,
        'review_rating': review_rating,
        'image_url': image_url
    }


def save_to_csv(book_info: Dict[str, str]) -> None:
    """Sauvegarde les informations du livre dans un fichier CSV nommé d'après le titre du livre.
    Args:
        book_info: Dictionnaire contenant les informations du livre.
    """
    title = book_info['title']
    safe_filename = re.sub(r'[^\w\s-]', '', title).strip().replace(' ', '_') + '.csv'

    fieldnames = [
        'product_page_url',
        'universal_product_code',
        'title',
        'price_including_tax',
        'price_excluding_tax',
        'number_available',
        'product_description',
        'category',
        'review_rating',
        'image_url'
    ]

    with open(safe_filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=';')
        writer.writeheader()
        writer.writerow(book_info)


def save_to_excel(book_info: Dict[str, str]) -> None:
    """Sauvegarde les informations du livre dans un fichier Excel nommé d'après le titre du livre, avec des colonnes colorées.
    Args:
        book_info: Dictionnaire contenant les informations du livre.
    """
    title = book_info['title']
    safe_filename = re.sub(r'[^\w\s-]', '', title).strip().replace(' ', '_') + '.xlsx'

    fieldnames = [
        'product_page_url',
        'universal_product_code',
        'title',
        'price_including_tax',
        'price_excluding_tax',
        'number_available',
        'product_description',
        'category',
        'review_rating',
        'image_url'
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Book Data"

    for col_idx, field in enumerate(fieldnames, start=1):
        ws.cell(row=1, column=col_idx).value = field

    for col_idx, field in enumerate(fieldnames, start=1):
        ws.cell(row=2, column=col_idx).value = book_info[field]

    colors = [
        'FFCC99', 'CCFFCC', '99CCFF', 'FFCCCC', 'FFFFCC',
        'CC99FF', '99FFFF', 'FF99CC', 'CCCCFF', '99FF99'
    ]

    for col_idx, color in enumerate(colors, start=1):
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        column_letter = get_column_letter(col_idx)
        ws[f'{column_letter}1'].fill = fill

    for col_idx in range(1, len(fieldnames) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    wb.save(safe_filename)


async def get_book_urls_from_category(category_url: str, session: aiohttp.ClientSession) -> List[str]:
    """Récupère les URLs de tous les livres d'une catégorie, en gérant la pagination (asynchrone).
    Args:
        category_url: L'URL de la page de la catégorie.
        session: Session aiohttp pour les requêtes.
    Returns:
        Une liste contenant les URLs de tous les livres de la catégorie.
    """
    book_urls = []
    current_url = category_url
    semaphore = asyncio.Semaphore(20)  # 20 requêtes simultanées

    while True:
        async with semaphore:
            try:
                soup = await async_fetch_page(session, current_url)
                books = soup.find_all('article', class_='product_pod')
                for book in books:
                    book_relative_url = book.find('h3').find('a')['href']
                    book_url = urljoin(current_url, book_relative_url)
                    book_urls.append(book_url)

                next_page = soup.find('li', class_='next')
                if next_page:
                    next_page_url = next_page.find('a')['href']
                    current_url = urljoin(current_url, next_page_url)
                else:
                    break

            except aiohttp.ClientError as e:
                print(f"Erreur lors de la récupération de la page {current_url} : {e}")
                break

    return book_urls


async def scrape_category(category_url: str) -> List[Dict[str, str]]:
    """Scrape les données de tous les livres d'une catégorie (asynchrone).
    Args:
        category_url: L'URL de la page de la catégorie.
    Returns:
        Une liste de dictionnaires, chaque dictionnaire contenant les informations d'un livre.
    """
    async with aiohttp.ClientSession() as session:
        book_urls = await get_book_urls_from_category(category_url, session)
        books_data = []
        semaphore = asyncio.Semaphore(20)  # 20 requêtes simultanées

        async def scrape_book(book_url: str) -> dict[str, str] | None:
            async with semaphore:
                try:
                    soup = await async_fetch_page(session, book_url)
                    book_info = extract_book_info(soup, book_url)
                    print(f"Livre scrapé : {book_info['title']}")
                    return book_info
                except (aiohttp.ClientError, AttributeError) as e:
                    print(f"Erreur lors du scraping du livre {book_url} : {e}")
                    return None

        tasks = [scrape_book(url) for url in book_urls]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        for result in results:
            if result is not None:
                books_data.append(result)

    return books_data


def save_category_to_csv(books_data: List[Dict[str, str]], category_name: str) -> None:
    """Sauvegarde les données de tous les livres d'une catégorie dans un fichier CSV.
    Args:
        books_data: Liste de dictionnaires contenant les informations des livres.
        category_name: Nom de la catégorie pour nommer le fichier.
    """
    safe_filename = re.sub(r'[^\w\s-]', '', category_name).strip().replace(' ', '_') + '.csv'

    fieldnames = [
        'product_page_url',
        'universal_product_code',
        'title',
        'price_including_tax',
        'price_excluding_tax',
        'number_available',
        'product_description',
        'category',
        'review_rating',
        'image_url'
    ]

    with open(safe_filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=';')
        writer.writeheader()
        for book_info in books_data:
            writer.writerow(book_info)


def generate_random_string(length: int = 8) -> str:
    """Génère une chaîne aléatoire de lettres minuscules et chiffres.
    Args:
        length: Longueur de la chaîne (par défaut 8).
    Returns:
        Une chaîne aléatoire de la longueur spécifiée.
    """
    characters = string.ascii_lowercase + string.digits
    return ''.join(random.choice(characters) for _ in range(length))


async def async_download_book_images(books_data: List[Dict[str, str]], image_dir: str = "images") -> None:
    """Télécharge les images de couverture de tous les livres dans le dossier spécifié (asynchrone).
    Args:
        books_data: Liste de dictionnaires contenant les informations des livres.
        image_dir: Dossier où sauvegarder les images (par défaut 'images').
    """
    os.makedirs(image_dir, exist_ok=True)
    semaphore = asyncio.Semaphore(20)  # 20 téléchargements simultanés

    async def download_image(session: aiohttp.ClientSession, book_info: Dict[str, str]) -> None:
        async with semaphore:
            try:
                image_url = book_info['image_url']
                category = book_info['category']
                safe_category = re.sub(r'[^\w\s-]', '', category).strip().replace(' ', '_').lower()
                random_str = generate_random_string()
                filename = f"{safe_category}_{random_str}.png"
                filepath = os.path.join(image_dir, filename)

                async with session.get(image_url) as response:
                    response.raise_for_status()
                    with open(filepath, 'wb') as f:
                        f.write(await response.read())
                print(f"Image téléchargée : {filename}")

            except (aiohttp.ClientError, IOError) as e:
                print(f"Erreur lors du téléchargement de l'image pour {book_info['title']} : {e}")

    async with aiohttp.ClientSession() as session:
        tasks = [download_image(session, book_info) for book_info in books_data]
        await asyncio.gather(*tasks)


def generate_category_summary_csv(books_data: List[Dict[str, str]],
                                  filename: str = "books_details_by_category.csv") -> None:
    """Génère un fichier CSV avec le nombre de livres et le prix moyen par catégorie.
    Args:
        books_data: Liste de dictionnaires contenant les informations des livres.
        filename: Nom du fichier CSV (par défaut 'books_details_by_category.csv').
    """
    # DataFrame
    df = pd.DataFrame(books_data)

    # Nettoyer et convertir les prix en float (en supprimant le symbole £)
    df['price_including_tax'] = df['price_including_tax'].str.replace('£', '').astype(float)

    # grouper par catégorie et calculer le nombre de livres et le prix moyen
    summary = df.groupby('category').agg({
        'title': 'count',  # nombre de livres
        'price_including_tax': 'mean'  # prix moyen
    }).rename(columns={'title': 'books_count', 'price_including_tax': 'average_price'}).reset_index()

    # prix moyen à 2 décimales
    summary['average_price'] = summary['average_price'].round(2)

    # fichier CSV avec délimiteur point-virgule
    summary.to_csv(filename, sep=';', index=False, encoding='utf-8')
    print(f"Fichier CSV '{filename}' généré avec succès.")


def plot_pie_chart(books_data: List[Dict[str, str]], output_file: str = "books_by_category_pie_chart.png") -> None:
    """Génère un diagramme circulaire montrant la répartition des livres pour les 20 premières catégories.
    Args:
        books_data: Liste de dictionnaires contenant les informations des livres.
        output_file: Nom du fichier image pour sauvegarder le diagramme (par défaut 'books_by_category_pie_chart.png').
    """
    df = pd.DataFrame(books_data)

    # nombre de livres par catégorie et trier par ordre décroissant
    category_counts = df['category'].value_counts().head(20)  # 20 premières catégories

    # diagramme circulaire
    plt.figure(figsize=(12, 8))
    plt.pie(category_counts, labels=category_counts.index, autopct='%1.1f%%', startangle=140,
            colors=plt.cm.Paired.colors)
    plt.title("Répartition des livres par catégorie (Top 20)")
    plt.axis('equal')  # Assurer que le diagramme est circulaire

    # save
    plt.savefig(output_file, bbox_inches='tight')
    print(f"Diagramme circulaire sauvegardé dans '{output_file}'.")

    plt.show()


def plot_price_histogram(books_data: List[Dict[str, str]],output_file: str = "average_price_by_category_histogram.png") -> None:
    """Génère un histogramme montrant les prix moyens des livres par catégorie.
    Args:
        books_data: Liste de dictionnaires contenant les informations des livres.
        output_file: Nom du fichier image pour sauvegarder l'histogramme (par défaut 'average_price_by_category_histogram.png').
    """
    df = pd.DataFrame(books_data)

    df['price_including_tax'] = df['price_including_tax'].str.replace('£', '').astype(float)

    # catégorie et calculer le prix moyen
    summary = df.groupby('category')['price_including_tax'].mean().sort_values()

    # histogramme
    plt.figure(figsize=(15, 8))
    summary.plot(kind='bar', color='skyblue')
    plt.title("Prix moyen des livres par catégorie")
    plt.xlabel("Catégorie")
    plt.ylabel("Prix moyen (£)")
    plt.xticks(rotation=45, ha='right')
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    plt.tight_layout()

    plt.savefig(output_file, bbox_inches='tight')
    print(f"Histogramme sauvegardé dans '{output_file}'.")

    plt.show()


def main() -> None:
    """Fonction principale pour exécuter le scraping des étapes 1, 2, 3, 4 et bonus."""
    global all_books_data
    start_time = time.time()  # Enregistrer le temps de début

    # Étape 1 : Scraper un seul livre et sauvegarder dans un fichier Excel (synchrone)
    print("=== Étape 1 : Scraping d'un seul livre ===")
    book_url = "http://books.toscrape.com/catalogue/a-light-in-the-attic_1000/index.html"
    try:
        soup = fetch_page(book_url)
        book_info = extract_book_info(soup, book_url)
        save_to_excel(book_info)
        print(f"Données sauvegardées dans '{book_info['title'].replace(' ', '_')}.xlsx'")
    except requests.RequestException as e:
        print(f"Erreur lors de la requête HTTP : {e}")
    except AttributeError as e:
        print(f"Erreur lors de l'extraction des données : {e}")
    except IOError as e:
        print(f"Erreur lors de l'écriture du fichier : {e}")

    # Étape 2 : Scraper tous les livres d'une catégorie et sauvegarder dans un fichier CSV (asynchrone)
    print("\n=== Étape 2 : Scraping d'une catégorie entière ===")
    category = "travel_2"
    category_url = "https://books.toscrape.com/catalogue/category/books/" + category + "/index.html"
    category_name = category[:-2].capitalize()
    try:
        books_data = asyncio.run(scrape_category(category_url))
        if books_data:
            save_category_to_csv(books_data, category_name)
            print(f"Données de la catégorie sauvegardées dans '{category_name.replace(' ', '_')}.csv'")
        else:
            print("Aucun livre trouvé dans la catégorie.")
    except Exception as e:
        print(f"Erreur lors du scraping de la catégorie : {e}")

    # Étape 3 : Scraper tous les livres du site (catégorie 'Books') et sauvegarder dans un fichier CSV (asynchrone)
    print("\n=== Étape 3 : Scraping de tous les livres ===")
    all_books_url = "https://books.toscrape.com/catalogue/category/books_1/index.html"
    all_books_category_name = "All_Books"
    try:
        all_books_data = asyncio.run(scrape_category(all_books_url))
        if all_books_data:
            save_category_to_csv(all_books_data, all_books_category_name)
            print(f"Données de tous les livres sauvegardées dans '{all_books_category_name.replace(' ', '_')}.csv'")
        else:
            print("Aucun livre trouvé dans la catégorie 'Books'.")
    except Exception as e:
        print(f"Erreur lors du scraping de tous les livres : {e}")

    # Étape 4 : Télécharger les images de tous les livres (asynchrone)
    print("\n=== Étape 4 : Téléchargement des images de tous les livres ===")
    try:
        if all_books_data:
            asyncio.run(async_download_book_images(all_books_data))
            print(f"Images téléchargées dans le dossier 'images'")
        else:
            print("Aucune donnée de livre disponible pour télécharger les images.")
    except Exception as e:
        print(f"Erreur lors du téléchargement des images : {e}")

    # Étape Bonus 1 : Générer le CSV de résumé par catégorie et le diagramme circulaire (Top 20 catégories)
    print("\n=== Étape Bonus 1 : Génération du CSV et diagramme circulaire (Top 20 catégories) ===")
    try:
        if all_books_data:
            generate_category_summary_csv(all_books_data)
            plot_pie_chart(all_books_data)
        else:
            print("Aucune donnée de livre disponible pour générer le CSV et le diagramme.")
    except Exception as e:
        print(f"Erreur lors de la génération du CSV ou du diagramme circulaire : {e}")

    # Étape Bonus 2 : Générer l'histogramme des prix moyens par catégorie
    print("\n=== Étape Bonus 2 : Génération de l'histogramme des prix moyens par catégorie ===")
    try:
        if all_books_data:
            plot_price_histogram(all_books_data)
        else:
            print("Aucune donnée de livre disponible pour générer l'histogramme.")
    except Exception as e:
        print(f"Erreur lors de la génération de l'histogramme : {e}")

    # Calculer et afficher le temps d'exécution total
    end_time = time.time()
    execution_time = end_time - start_time
    minutes = int(execution_time // 60)
    seconds = int(execution_time % 60)
    print(f"\nTemps d'exécution total : {minutes} minutes et {seconds} secondes ({execution_time:.2f} secondes)")


if __name__ == "__main__":
    main()
